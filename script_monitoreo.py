from openpyxl import load_workbook , Workbook
from datetime import datetime
from icmplib import ping 



"""Leemos el Archivo de Excel que contiene la lista de IP's a Monitorear"""

file_load = load_workbook("Lista_IP.xlsx")
sheet_file = file_load['Listado']

data = []  #variable para almacenar los datos
"""Se itera las filas del archivo que contienen valores y los datos los agregamos a la variable {data}
"""
for row in sheet_file.iter_rows(min_col=1,min_row=1, max_col=2, values_only=True):
    data.append(row)

#convertimos la informacion en un diccionario y borramos su primer elemento
data_end = dict(data) 
del data_end['Equipo | Descripcion'] 

"""Creamos el archivo Excel(xlsx) que contendra los resultados"""

date = datetime.now()
name_file = f"Report_{date.day}{date.month}{date.year}_{date.hour}{date.minute}"
wb = Workbook()
dest_filename=f"{name_file}.xlsx"
file = wb.active
file.title = "Results"
file.append(('Equipo | Descripcion', 
             'IP', 
             'Time Average Ping', 
             'Packet Sent' , 
             'Packet Received', 
             'Latency'))
data_upload = []

print("Empezando a Realizar la primera Validacion del Ping")
TEST_START =datetime.now()
print(TEST_START)


for id, ip in data_end.items():
    response_host = ping(ip, count=4 , interval=1)
    time_avrg = response_host.avg_rtt
    pkt_sent = response_host.packets_sent
    pkt_received = response_host.packets_received
    latency = response_host.jitter
    
    data_upload.append((f'{id}', 
                        f'{ip}', 
                        f'{time_avrg}', 
                        f'{pkt_sent}', 
                        f'{pkt_received}', 
                        f'{latency}'))

for data_file in data_upload:
    file.append(data_file)

wb.save(filename= dest_filename)
TEST_END =datetime.now()
print(TEST_END)
print("Test Finalizado")
print(TEST_END-TEST_START)

    
    
    
    
    








